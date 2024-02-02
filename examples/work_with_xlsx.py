import os
from openpyxl import load_workbook

PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))

# with open(os.path.join(PROJECT_ROOT_PATH, "resources/template.xls"), 'r') as xls:
#     reader = open_workbook(xls)
#     print(reader.sheet_names())

workbook = load_workbook(os.path.join(PROJECT_ROOT_PATH, "resources/calendare.xlsx"))
sheet = workbook.active

print(sheet.cell(3, 3).value)
