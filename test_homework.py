import os
import zipfile
from zipfile import ZipFile
import PyPDF2
from openpyxl import load_workbook
from xlrd import open_workbook
import csv


def create_zip(file_archive):
    path = os.path.join(os.getcwd(), 'resources')
    list_dir = os.listdir(path)
    destination = os.path.join(path, file_archive)
    with ZipFile(destination, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file_name in list_dir:
            add_file = os.path.join(path, file_name)
            zf.write(add_file, os.path.basename(add_file))


def add_file_to_zip(name_archive, name_file):
    destination = os.path.join(os.getcwd(), 'resources', name_archive)
    with ZipFile(destination, 'a', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(name_file)


def list_files_to_zip(name_archive):
    destination = os.path.join(os.getcwd(), 'resources', name_archive)
    with ZipFile(destination, 'a') as zf:
        print(zf.namelist())


def extract_all_files(name_archive):
    extract_dir = 'extract_dir'
    destination = os.path.join(os.getcwd(), 'resources', name_archive)
    with ZipFile(destination) as zf:
        zf.extractall(extract_dir)


def extract_file(name_archive, name_file):
    extract_dir = 'extract_dir_one'
    destination = os.path.join(os.getcwd(), 'resources', name_archive)
    with ZipFile(destination) as zf:
        zf.extract(name_file, extract_dir)


create_zip('test1.zip')
add_file_to_zip('test1.zip', 'README.md')
list_files_to_zip('test1.zip')
extract_all_files('test1.zip')
extract_file('test1.zip', 'README.rst')


def test_pdf():
    pdf_file = 'english_grammar_book_for_students.pdf'
    destination = os.path.join(os.getcwd(), 'resources')
    with open(os.path.join(destination, pdf_file), 'rb') as pdf:
        reader = PyPDF2.PdfReader(pdf)
        text = reader.pages[2].extract_text()
        assert 'SERIES CONSULTANT' in text


def test_xlsx():
    xlsx_file = 'calendare.xlsx'
    destination = os.path.join(os.getcwd(), 'resources')
    workbook = load_workbook(os.path.join(destination, xlsx_file))
    sheet = workbook.active
    value = sheet.cell(3, 3).value
    assert 'Введение в компьютерные приложения' in value


def test_xls():
    xls_file = 'template.xls'
    destination = os.path.join(os.getcwd(), 'resources')
    reader = open_workbook(os.path.join(destination, xls_file))
    value = reader.sheet_by_index(0).cell_value(4, 1)
    assert 'Verify the is correct' in value


def test_csv():
    csv_file = 'csv_test.csv'
    destination = os.path.join(os.getcwd(), 'resources')
    with open(os.path.join(destination, csv_file)) as f:
        reader = csv.reader(f)
        value = reader.__next__()
    assert 'Value 1' in value


def test_txt():
    txt_file = 'README.rst'
    destination = os.path.join(os.getcwd(), 'resources')
    with open(os.path.join(destination, txt_file)) as txt:
        reader = txt.read()
        text = reader
    assert 'align: center' in text
